from flask import Flask, render_template_string, request, redirect, session, send_file
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo
from supabase_config import supabase

app = Flask(__name__)
app.secret_key = "clave_secreta_sistema_somnolencia"

BUCKET_EVIDENCIAS = "evidencias"

def convertir_fecha_peru(fecha_iso):
    if not fecha_iso:
        return ""

    try:
        fecha = datetime.fromisoformat(fecha_iso.replace("Z", "+00:00"))
        fecha_peru = fecha.astimezone(ZoneInfo("America/Lima"))
        return fecha_peru.strftime("%d/%m/%Y %H:%M:%S")
    except:
        return fecha_iso

def login_requerido():
    return "admin" in session

def cargar_alertas_por_dni(dni):
    return (
        supabase
        .table("alertas")
        .select("*")
        .eq("dni", dni)
        .order("id")
        .execute()
        .data
    )
    
def listar_imagenes_supabase(dni):
    try:
        archivos = (
            supabase
            .storage
            .from_(BUCKET_EVIDENCIAS)
            .list(dni)
        )

        imagenes = []

        for archivo in archivos:
            nombre = archivo.get("name", "")

            if nombre.lower().endswith((".jpg", ".jpeg", ".png")):
                ruta = f"{dni}/{nombre}"

                url = (
                    supabase
                    .storage
                    .from_(BUCKET_EVIDENCIAS)
                    .get_public_url(ruta)
                )

                imagenes.append({
                    "nombre": nombre,
                    "ruta": ruta,
                    "url": url
                })

        return imagenes

    except Exception as e:
        print("Error listando imágenes:", e)
        return []


def borrar_imagen_supabase(ruta):
    try:
        supabase.storage.from_(BUCKET_EVIDENCIAS).remove([ruta])
    except Exception as e:
        print("Error borrando imagen:", e)


def borrar_imagenes_supabase(dni):
    imagenes = listar_imagenes_supabase(dni)

    rutas = [
        img["ruta"]
        for img in imagenes
    ]

    if rutas:
        try:
            supabase.storage.from_(BUCKET_EVIDENCIAS).remove(rutas)
        except Exception as e:
            print("Error borrando imágenes:", e)

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"]
        password = request.form["password"]

        admins = (
            supabase
            .table("admins")
            .select("*")
            .execute()
            .data
        )

        for admin in admins:
            if admin["usuario"] == usuario and admin["password"] == password:
                session["admin"] = usuario
                return redirect("/dashboard")

        return render_template_string(LOGIN_HTML, error="Usuario o contraseña incorrectos")

    return render_template_string(LOGIN_HTML, error=None)


@app.route("/dashboard")
def dashboard():
    if not login_requerido():
        return redirect("/")

    conductores = (
        supabase
        .table("conductores")
        .select("*")
        .order("id")
        .execute()
        .data
    )
    for c in conductores:
        c["fecha_registro"] = convertir_fecha_peru(c.get("fecha_registro"))
    return render_template_string(DASHBOARD_HTML, conductores=conductores, admin=session["admin"])


@app.route("/evidencias/<dni>")
def evidencias(dni):
    if not login_requerido():
        return redirect("/")

    conductores = (
        supabase
        .table("conductores")
        .select("*")
        .execute()
        .data
    )
    alertas_conductor = cargar_alertas_por_dni(dni)

    conductor = None
    for c in conductores:
        if c["dni"] == dni:
            conductor = c
            break

    imagenes = listar_imagenes_supabase(dni)

    return render_template_string(
        EVIDENCIAS_HTML,
        conductor=conductor,
        dni=dni,
        imagenes=imagenes,
        alertas=alertas_conductor,
        admin=session["admin"]
    )


@app.route("/borrar-alerta/<dni>/<id_alerta>", methods=["POST"])
def borrar_alerta(dni, id_alerta):
    if not login_requerido():
        return redirect("/")

    supabase \
        .table("alertas") \
        .delete() \
        .eq("id", id_alerta) \
        .execute()

    return redirect(f"/evidencias/{dni}")


@app.route("/borrar-alertas/<dni>", methods=["POST"])
def borrar_alertas(dni):
    if not login_requerido():
        return redirect("/")

    supabase \
        .table("alertas") \
        .delete() \
        .eq("dni", dni) \
        .execute()

    return redirect(f"/evidencias/{dni}")


@app.route("/borrar-imagen/<dni>/<nombre>", methods=["POST"])
def borrar_imagen(dni, nombre):
    if not login_requerido():
        return redirect("/")

    ruta = f"{dni}/{nombre}"
    borrar_imagen_supabase(ruta)

    return redirect(f"/evidencias/{dni}")


@app.route("/borrar-conductor/<dni>", methods=["POST"])
def borrar_conductor(dni):

    if not login_requerido():
        return redirect("/")

    # Eliminar conductor de SUPABASE
    try: 
        supabase \
            .table("conductores") \
            .delete() \
            .eq("dni", dni) \
            .execute()
    
    except Exception as e:
        print("Error eliminando conductor:", e)

    # Eliminar alertas 
    # Eliminar alertas de Supabase
    supabase \
        .table("alertas") \
        .delete() \
        .eq("dni", dni) \
        .execute()

    # Eliminar imágenes de Supabase Storage
    borrar_imagenes_supabase(dni)# Eliminar imágenes

    return redirect("/dashboard")

@app.route("/borrar-imagenes/<dni>", methods=["POST"])
def borrar_imagenes(dni):
    if not login_requerido():
        return redirect("/")

    borrar_imagenes_supabase(dni)

    return redirect(f"/evidencias/{dni}")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


LOGIN_HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Login Administrador</title>
    <style>
        body {
            font-family: Arial;
            background: #111827;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            margin: 0;
        }

        .login-box {
            background: white;
            padding: 35px;
            border-radius: 12px;
            width: 350px;
            box-shadow: 0 8px 20px rgba(0,0,0,0.3);
            text-align: center;
        }

        input {
            width: 90%;
            padding: 12px;
            margin: 10px;
            border-radius: 6px;
            border: 1px solid #ccc;
        }

        button {
            width: 95%;
            padding: 12px;
            background: #2563eb;
            color: white;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-weight: bold;
        }

        .error {
            color: red;
            margin-top: 10px;
        }
    </style>
</head>
<body>
    <div class="login-box">
        <h2>Login Administrador</h2>
        <p>Sistema de Somnolencia</p>

        <form method="POST">
            <input type="text" name="usuario" placeholder="Usuario" required>
            <input type="password" name="password" placeholder="Contraseña" required>
            <button type="submit">Ingresar</button>
        </form>

        {% if error %}
            <p class="error">{{ error }}</p>
        {% endif %}
    </div>
</body>
</html>
"""


DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Dashboard</title>
    <style>
    
        .danger-small {
            background: #dc2626;
            color: white;
            border: none;
            padding: 8px 10px;
            border-radius: 6px;
            cursor: pointer;
            margin-left: 5px;
        }
    
        body {
            font-family: Arial;
            background: #f4f6f8;
            margin: 0;
        }

        header {
            background: #111827;
            color: white;
            padding: 20px;
            text-align: center;
        }

        .top {
            text-align: right;
            padding: 15px;
        }

        .top a {
            background: #dc2626;
            color: white;
            padding: 8px 12px;
            text-decoration: none;
            border-radius: 6px;
            margin-left: 8px;
        }

        .reload {
            background: #16a34a !important;
        }

        .container {
            width: 90%;
            margin: 20px auto;
        }

        .card {
            background: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 4px 10px rgba(0,0,0,0.1);
        }

        table {
            width: 100%;
            border-collapse: collapse;
            background: white;
        }

        th, td {
            padding: 12px;
            border-bottom: 1px solid #ddd;
            text-align: center;
        }

        th {
            background: #1f2937;
            color: white;
        }

        .btn {
            background: #2563eb;
            color: white;
            padding: 8px 12px;
            text-decoration: none;
            border-radius: 6px;
        }
    </style>
</head>
<body>
    <header>
        <h1>Dashboard del Sistema Inteligente</h1>
        <p>Administrador: {{ admin }}</p>
    </header>

    <div class="top">
        <a class="reload" href="/dashboard">Recargar página</a>
        <a href="/logout">Cerrar sesión</a>
    </div>

    <div class="container">
        <div class="card">
            <h2>Conductores registrados</h2>

            <table>
                <tr>
                    <th>Nombre</th>
                    <th>DNI</th>
                    <th>Correo</th>
                    <th>Teléfono</th>
                    <th>Fecha registro</th>
                    <th>Acciones</th>
                </tr>

                {% for c in conductores %}
                <tr>
                    <td>{{ c.nombre }}</td>
                    <td>{{ c.dni }}</td>
                    <td>{{ c.correo }}</td>
                    <td>{{ c.telefono }}</td>
                    <td>{{ c.fecha_registro }}</td>
                    <td>

                        <a class="btn" href="/evidencias/{{ c.dni }}">
                            Ver evidencias
                        </a>

                        <form
                            method="POST"
                            action="/borrar-conductor/{{ c.dni }}"
                            style="display:inline;"
                            onsubmit="return confirm(
                            '¿Eliminar conductor, historial e imágenes?'
                            )">

                            <button
                                class="danger-small"
                                type="submit">

                                Eliminar

                            </button>

                        </form>

                    </td>
                </tr>
                {% endfor %}
            </table>
        </div>
    </div>
</body>
</html>
"""


EVIDENCIAS_HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Evidencias</title>
    <style>
        body {
            font-family: Arial;
            background: #f4f6f8;
            margin: 0;
        }

        header {
            background: #111827;
            color: white;
            padding: 20px;
            text-align: center;
        }

        .container {
            width: 92%;
            margin: 25px auto;
        }

        .card {
            background: white;
            padding: 20px;
            margin-bottom: 20px;
            border-radius: 10px;
            box-shadow: 0 4px 10px rgba(0,0,0,0.1);
        }

        .acciones {
            margin-bottom: 20px;
        }

        a {
            background: #2563eb;
            color: white;
            padding: 8px 12px;
            text-decoration: none;
            border-radius: 6px;
            margin-right: 8px;
        }

        .reload {
            background: #16a34a;
        }

        .danger {
            background: #dc2626;
            color: white;
            border: none;
            padding: 8px 12px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: bold;
        }

        .danger-small {
            background: #dc2626;
            color: white;
            border: none;
            padding: 6px 10px;
            border-radius: 6px;
            cursor: pointer;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            background: white;
        }

        th, td {
            padding: 12px;
            border-bottom: 1px solid #ddd;
            text-align: center;
        }

        th {
            background: #1f2937;
            color: white;
        }

        .grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(290px, 1fr));
            gap: 20px;
        }

        .img-card {
            background: #f9fafb;
            border-radius: 10px;
            padding: 12px;
            text-align: center;
            box-shadow: 0 3px 8px rgba(0,0,0,0.1);
        }

        img {
            width: 260px;
            border-radius: 10px;
            margin-bottom: 10px;
        }

        .nombre-img {
            font-size: 13px;
            word-break: break-all;
            margin-bottom: 10px;
            color: #111827;
            font-weight: bold;
        }

        form {
            display: inline;
        }
    </style>
</head>
<body>
    <header>
        <h1>Evidencias del Conductor</h1>
        {% if conductor %}
            <p>{{ conductor.nombre }} - DNI: {{ conductor.dni }}</p>
        {% endif %}
    </header>

    <div class="container">
        <div class="acciones">
            <a href="/dashboard">Volver</a>
            <a class="reload" href="/evidencias/{{ dni }}">Recargar evidencias</a>
            <a class="reload" href="/exportar-excel/{{ dni }}">Exportar Excel</a>
        </div>

        <div class="card">
            <h2>Historial de alertas</h2>

            <form method="POST" action="/borrar-alertas/{{ dni }}" onsubmit="return confirm('¿Seguro que deseas borrar TODO el historial de este conductor?')">
                <button class="danger" type="submit">Borrar todo el historial</button>
            </form>

            <br><br>

            <table>
                <tr>
                    <th>ID</th>
                    <th>Fecha</th>
                    <th>Hora</th>
                    <th>Tipo alerta</th>
                    <th>Nivel cansancio</th>
                    <th>Acción</th>
                </tr>

                {% for a in alertas %}
                <tr>
                    <td>{{ loop.index }}</td>
                    <td>{{ a.fecha }}</td>
                    <td>{{ a.hora }}</td>
                    <td>{{ a.tipo_alerta }}</td>
                    <td>{{ a.nivel_cansancio }}%</td>
                    <td>
                        <form method="POST" action="/borrar-alerta/{{ dni }}/{{ a.id }}" onsubmit="return confirm('¿Eliminar solo esta alerta?')">
                            <button class="danger-small" type="submit">Eliminar</button>
                        </form>
                    </td>
                </tr>
                {% endfor %}
            </table>
        </div>

        <div class="card">
            <h2>Imágenes de evidencia</h2>

            <form method="POST" action="/borrar-imagenes/{{ dni }}" onsubmit="return confirm('¿Seguro que deseas borrar TODAS las imágenes de este conductor?')">
                <button class="danger" type="submit">Borrar todas las imágenes</button>
            </form>

            <br><br>

            {% if imagenes %}
                <div class="grid">
                    {% for img in imagenes %}
                    <div class="img-card">
                        <img src="{{ img.url }}">
                        <div class="nombre-img">{{ img.nombre }}</div>

                        <form method="POST" action="/borrar-imagen/{{ dni }}/{{ img.nombre }}" onsubmit="return confirm('¿Eliminar esta imagen?')">
                            <button class="danger-small" type="submit">Eliminar imagen</button>
                        </form>
                    </div>
                    {% endfor %}
                </div>
            {% else %}
                <p>No hay imágenes registradas para este conductor.</p>
            {% endif %}
        </div>
    </div>
</body>
</html>
"""
@app.route("/api/registrar-conductor", methods=["POST"])
def api_registrar_conductor():
    data = request.json

    nombre = data.get("nombre")
    dni = data.get("dni")
    correo = data.get("correo")
    telefono = data.get("telefono")

    if not nombre or not dni or not correo or not telefono:
        return {"ok": False, "mensaje": "Faltan datos"}, 400

    existe = (
        supabase
        .table("conductores")
        .select("*")
        .or_(f"dni.eq.{dni},correo.eq.{correo}")
        .execute()
        .data
    )

    if existe:
        return {"ok": False, "mensaje": "El conductor ya existe"}, 400

    supabase.table("conductores").insert({
        "nombre": nombre,
        "dni": dni,
        "correo": correo,
        "telefono": telefono
    }).execute()

    return {"ok": True, "mensaje": "Conductor registrado correctamente"}


@app.route("/api/login-conductor", methods=["POST"])
def api_login_conductor():
    data = request.json

    correo = data.get("correo")
    dni = data.get("dni")

    conductor = (
        supabase
        .table("conductores")
        .select("*")
        .eq("correo", correo)
        .eq("dni", dni)
        .execute()
        .data
    )

    if not conductor:
        return {"ok": False, "mensaje": "Correo o DNI incorrecto"}, 401

    return {
        "ok": True,
        "conductor": conductor[0]
    }


@app.route("/api/registrar-alerta", methods=["POST"])
def api_registrar_alerta():
    data = request.json

    dni = data.get("dni")
    conductor = data.get("conductor")
    tipo_alerta = data.get("tipo_alerta")
    nivel_cansancio = data.get("nivel_cansancio")

    ahora = datetime.now()

    supabase.table("alertas").insert({
        "fecha": ahora.strftime("%d/%m/%Y"),
        "hora": ahora.strftime("%H:%M:%S"),
        "dni": dni,
        "conductor": conductor,
        "tipo_alerta": tipo_alerta,
        "nivel_cansancio": nivel_cansancio
    }).execute()

    return {"ok": True, "mensaje": "Alerta registrada"}


@app.route("/api/subir-evidencia", methods=["POST"])
def api_subir_evidencia():
    dni = request.form.get("dni")
    tipo_alerta = request.form.get("tipo_alerta")
    archivo = request.files.get("imagen")

    if not dni or not tipo_alerta or not archivo:
        return {"ok": False, "mensaje": "Faltan datos"}, 400

    nombre = f"{tipo_alerta}_{datetime.now().strftime('%d-%m-%Y_%H-%M-%S')}.jpg"
    ruta = f"{dni}/{nombre}"

    contenido = archivo.read()

    supabase.storage.from_(BUCKET_EVIDENCIAS).upload(
        ruta,
        contenido,
        {
            "content-type": "image/jpeg",
            "upsert": "true"
        }
    )

    return {
        "ok": True,
        "mensaje": "Evidencia subida",
        "ruta": ruta
    }

@app.route("/exportar-excel/<dni>")
def exportar_excel(dni):
    if not login_requerido():
        return redirect("/")

    conductores = (
        supabase
        .table("conductores")
        .select("*")
        .eq("dni", dni)
        .execute()
        .data
    )

    if not conductores:
        return "Conductor no encontrado", 404

    conductor = conductores[0]
    alertas = cargar_alertas_por_dni(dni)
    imagenes = listar_imagenes_supabase(dni)

    wb = Workbook()

    # Hoja 1: Datos del conductor
    ws1 = wb.active
    ws1.title = "Conductor"

    ws1.append(["Campo", "Valor"])
    ws1.append(["Nombre", conductor.get("nombre", "")])
    ws1.append(["DNI", conductor.get("dni", "")])
    ws1.append(["Correo", conductor.get("correo", "")])
    ws1.append(["Teléfono", conductor.get("telefono", "")])
    ws1.append(["Fecha registro", conductor.get("fecha_registro", "")])

    # Hoja 2: Alertas
    ws2 = wb.create_sheet("Alertas")
    ws2.append(["N°", "Fecha", "Hora", "Tipo de alerta", "Nivel cansancio"])

    for i, alerta in enumerate(alertas, start=1):
        ws2.append([
            i,
            alerta.get("fecha", ""),
            alerta.get("hora", ""),
            alerta.get("tipo_alerta", ""),
            f"{alerta.get('nivel_cansancio', '')}%"
        ])

    # Hoja 3: Evidencias
    ws3 = wb.create_sheet("Evidencias")
    ws3.append(["N°", "Nombre de imagen", "URL"])

    for i, img in enumerate(imagenes, start=1):
        ws3.append([
            i,
            img.get("nombre", ""),
            img.get("url", "")
        ])

    # Estilos simples
    for ws in [ws1, ws2, ws3]:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.alignment = Alignment(horizontal="center")

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(value))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 60)

    archivo_excel = BytesIO()
    wb.save(archivo_excel)
    archivo_excel.seek(0)

    nombre_archivo = f"reporte_conductor_{dni}.xlsx"

    return send_file(
        archivo_excel,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000
    )